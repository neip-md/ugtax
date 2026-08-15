"""Excel export for Buchungsjournal and Bilanz/GuV."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

from .models import FinancialStatements, JournalEntry


EUR_FORMAT = '#,##0.00 €'
HEADER_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content."""
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = min_width
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len


def export_journal(entries: list[JournalEntry], output_path: str | Path) -> None:
    """Export Buchungsjournal to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Buchungsjournal"

    # Headers
    headers = ["Nr.", "Datum", "Soll-Konto", "Soll-Bezeichnung", "Haben-Konto", "Haben-Bezeichnung", "Betrag (€)", "Beschreibung"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT

    # Data
    for i, entry in enumerate(entries, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=entry.date.isoformat())
        ws.cell(row=row, column=3, value=entry.debit_account)
        ws.cell(row=row, column=4, value=entry.debit_account_name)
        ws.cell(row=row, column=5, value=entry.credit_account)
        ws.cell(row=row, column=6, value=entry.credit_account_name)
        amt_cell = ws.cell(row=row, column=7, value=float(entry.amount))
        amt_cell.number_format = EUR_FORMAT
        ws.cell(row=row, column=8, value=entry.description).alignment = WRAP_ALIGNMENT

    # Totals row
    total_row = len(entries) + 2
    ws.cell(row=total_row, column=6, value="Summe:").font = HEADER_FONT
    total_cell = ws.cell(row=total_row, column=7, value=float(sum(e.amount for e in entries)))
    total_cell.number_format = EUR_FORMAT
    total_cell.font = HEADER_FONT

    _auto_width(ws)
    wb.save(output_path)


def export_statements(statements: FinancialStatements, output_path: str | Path) -> None:
    """Export Bilanz + GuV to Excel."""
    wb = Workbook()

    # Sheet 1: Bilanz
    ws_bilanz = wb.active
    ws_bilanz.title = "Bilanz"
    _write_bilanz(ws_bilanz, statements)

    # Sheet 2: GuV
    ws_guv = wb.create_sheet("GuV")
    _write_guv(ws_guv, statements)

    # Sheet 3: Kontensalden
    ws_konten = wb.create_sheet("Kontensalden")
    _write_konten(ws_konten, statements)

    wb.save(output_path)


def _write_bilanz(ws, statements: FinancialStatements) -> None:
    """Write Bilanz sheet."""
    bilanz = statements.bilanz

    # Title
    ws.cell(row=1, column=1, value=f"Bilanz zum 31.12.{statements.fiscal_year}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=statements.company_name)
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    row = 4
    # Aktiva
    ws.cell(row=row, column=1, value="AKTIVA").font = HEADER_FONT
    ws.cell(row=row, column=2, value="EUR").font = HEADER_FONT
    row += 1

    for position, amount in sorted(bilanz.aktiva.items()):
        if amount != Decimal("0.00"):
            ws.cell(row=row, column=1, value=position)
            cell = ws.cell(row=row, column=2, value=float(amount))
            cell.number_format = EUR_FORMAT
            row += 1

    ws.cell(row=row, column=1, value="Summe Aktiva").font = HEADER_FONT
    total = ws.cell(row=row, column=2, value=float(bilanz.summe_aktiva))
    total.number_format = EUR_FORMAT
    total.font = HEADER_FONT
    row += 2

    # Passiva
    ws.cell(row=row, column=1, value="PASSIVA").font = HEADER_FONT
    ws.cell(row=row, column=2, value="EUR").font = HEADER_FONT
    row += 1

    for position, amount in sorted(bilanz.passiva.items()):
        if amount != Decimal("0.00"):
            ws.cell(row=row, column=1, value=position)
            cell = ws.cell(row=row, column=2, value=float(amount))
            cell.number_format = EUR_FORMAT
            row += 1

    ws.cell(row=row, column=1, value="Summe Passiva").font = HEADER_FONT
    total = ws.cell(row=row, column=2, value=float(bilanz.summe_passiva))
    total.number_format = EUR_FORMAT
    total.font = HEADER_FONT

    _auto_width(ws)


def _write_guv(ws, statements: FinancialStatements) -> None:
    """Write GuV sheet."""
    guv = statements.guv

    ws.cell(row=1, column=1, value=f"Gewinn- und Verlustrechnung {statements.fiscal_year}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=statements.company_name)
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    row = 4
    # Erträge
    ws.cell(row=row, column=1, value="ERTRÄGE").font = HEADER_FONT
    row += 1
    for position, amount in sorted(guv.ertraege.items()):
        if amount != Decimal("0.00"):
            ws.cell(row=row, column=1, value=position)
            cell = ws.cell(row=row, column=2, value=float(amount))
            cell.number_format = EUR_FORMAT
            row += 1

    if not guv.ertraege:
        ws.cell(row=row, column=1, value="(keine)")
        row += 1

    ws.cell(row=row, column=1, value="Summe Erträge").font = HEADER_FONT
    cell = ws.cell(row=row, column=2, value=float(guv.summe_ertraege))
    cell.number_format = EUR_FORMAT
    cell.font = HEADER_FONT
    row += 2

    # Aufwendungen
    ws.cell(row=row, column=1, value="AUFWENDUNGEN").font = HEADER_FONT
    row += 1
    for position, amount in sorted(guv.aufwendungen.items()):
        if amount != Decimal("0.00"):
            ws.cell(row=row, column=1, value=position)
            cell = ws.cell(row=row, column=2, value=float(amount))
            cell.number_format = EUR_FORMAT
            row += 1

    ws.cell(row=row, column=1, value="Summe Aufwendungen").font = HEADER_FONT
    cell = ws.cell(row=row, column=2, value=float(guv.summe_aufwendungen))
    cell.number_format = EUR_FORMAT
    cell.font = HEADER_FONT
    row += 2

    # Jahresüberschuss
    ws.cell(row=row, column=1, value="JAHRESÜBERSCHUSS / JAHRESFEHLBETRAG").font = Font(bold=True, size=12)
    cell = ws.cell(row=row, column=2, value=float(guv.jahresueberschuss))
    cell.number_format = EUR_FORMAT
    cell.font = Font(bold=True, size=12)

    _auto_width(ws)


def _write_konten(ws, statements: FinancialStatements) -> None:
    """Write Kontensalden sheet."""
    ws.cell(row=1, column=1, value="Kontensalden").font = Font(bold=True, size=14)

    headers = ["Konto", "Bezeichnung", "Kategorie", "Soll", "Haben", "Saldo"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header).font = HEADER_FONT

    row = 4
    for number in sorted(statements.account_balances.keys()):
        bal = statements.account_balances[number]
        if bal.debit_total == Decimal("0.00") and bal.credit_total == Decimal("0.00"):
            continue
        ws.cell(row=row, column=1, value=number)
        ws.cell(row=row, column=2, value=bal.name)
        ws.cell(row=row, column=3, value=bal.category.value)
        ws.cell(row=row, column=4, value=float(bal.debit_total)).number_format = EUR_FORMAT
        ws.cell(row=row, column=5, value=float(bal.credit_total)).number_format = EUR_FORMAT
        ws.cell(row=row, column=6, value=float(bal.balance)).number_format = EUR_FORMAT
        row += 1

    _auto_width(ws)
